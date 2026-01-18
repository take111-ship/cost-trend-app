# app.py
import io
import requests
import pandas as pd
import streamlit as st
import matplotlib.pyplot as plt
from datetime import date, datetime

# Excel export
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows

# Pillow is required by openpyxl for image handling
# requirements.txt: add Pillow
from PIL import Image as PILImage  # noqa: F401


# ----------------------------
# Page / UI
# ----------------------------
st.set_page_config(
    page_title="原価ダッシュボード（銅・アルミ・運賃・賃金）",
    page_icon="📈",
    layout="wide",
)

st.title("📈 原価ダッシュボード（銅・アルミ・運賃・賃金）")
st.caption("銅/アルミはFRED（USD/ton）×USDJPYで円/kg換算。")

with st.sidebar:
    st.header("📚 データ元リンク（固定表示）")
    st.markdown("""
**■ FRED**
- 銅（USD/ton）: https://fred.stlouisfed.org/series/PCOPPUSDM  
- アルミ（USD/ton）: https://fred.stlouisfed.org/series/PALUMUSDM  
- 為替（USD/JPY）: https://fred.stlouisfed.org/series/EXJPUS   
""")

with st.expander("このダッシュボードについて", expanded=True):
    st.markdown("""
- **銅・アルミ（円/kg）**：FREDの月次（USD/ton）× USDJPY ÷ 1000  
- **Excel出力**：表示中データを統合し、表＋グラフ画像を添付したExcelを生成  
""")


# ----------------------------
# Secrets
# ----------------------------
FRED_API_KEY = st.secrets.get("FRED_API_KEY", "")
if not FRED_API_KEY:
    st.error("FRED_API_KEY が設定されていません（Streamlit Secrets を確認）")
    st.stop()
# ----------------------------
# FRED fetch
# ----------------------------
@st.cache_data(ttl=60 * 60)
def fetch_fred(series_id: str, start: str = "2018-01-01") -> pd.Series:
    url = "https://api.stlouisfed.org/fred/series/observations"
    params = {
        "series_id": series_id,
        "api_key": FRED_API_KEY,
        "file_type": "json",
        "observation_start": start,
        "observation_end": date.today().isoformat(),
    }
    r = requests.get(url, params=params, timeout=30)
    r.raise_for_status()
    obs = r.json()["observations"]

    df = pd.DataFrame(obs)[["date", "value"]]
    df["date"] = pd.to_datetime(df["date"])
    df["value"] = pd.to_numeric(df["value"], errors="coerce")
    s = df.dropna().set_index("date")["value"].sort_index()
    return s


# ----------------------------
# Common plot helper
# ----------------------------
def plot_with_latest_highlight(series: pd.Series, title: str, y_label: str):
    if series is None or series.empty:
        st.info("データが空です。")
        return

    fig, ax = plt.subplots()
    ax.plot(series.index, series.values)
    ax.scatter(series.index[-1], series.values[-1], s=80, zorder=3)
    ax.annotate(
        f"{series.values[-1]:,.0f}",
        (series.index[-1], series.values[-1]),
        textcoords="offset points",
        xytext=(8, 8),
        ha="left",
    )
    ax.set_title(title)
    ax.set_xlabel("Month")
    ax.set_ylabel(y_label)
    ax.grid(True, alpha=0.3)
    st.pyplot(fig, clear_figure=True)


# ----------------------------
# Excel export helpers
# ----------------------------
def _df_to_sheet(ws, df: pd.DataFrame, start_row=1, start_col=1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, v in enumerate(row, start_col):
            ws.cell(row=r_idx, column=c_idx, value=v)


def _series_chart_png_bytes(series: pd.Series, title: str, y_label: str) -> bytes:
    fig, ax = plt.subplots(figsize=(10, 4))
    ax.plot(series.index, series.values)
    ax.scatter(series.index[-1], series.values[-1], s=120, zorder=3)
    ax.set_title(title)
    ax.set_xlabel("Month")
    ax.set_ylabel(y_label)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=200)
    plt.close(fig)
    return buf.getvalue()


def build_monthly_master(df_fred: pd.DataFrame) -> pd.DataFrame:
    out = df_fred.copy()
    out.index = pd.to_datetime(out.index)
    # 月次へ寄せて月初に統一（pandas互換：how="start"）
    out.index = out.index.to_period("M").to_timestamp(how="start")
    out.index.name = "month"
    return out.sort_index()


def make_excel_report(master: pd.DataFrame) -> bytes:
    if master is None or master.dropna(how="all").empty:
        raise ValueError("Excel出力対象のデータが空です（masterが空）。")

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_data = wb.create_sheet("Data")
    ws_charts = wb.create_sheet("Charts")

    # --- Summary
    base = master.dropna(how="all")
    latest_row = base.iloc[-1]
    latest_month = base.index[-1].strftime("%Y-%m")

    ws_sum["A1"] = "最新月"
    ws_sum["B1"] = latest_month
    ws_sum["A3"] = "指標"
    ws_sum["B3"] = "最新値"
    ws_sum["C3"] = "前月差"

    summary_rows = []
    for col in master.columns:
        s = master[col].dropna()
        if s.empty:
            continue
        v_now = float(s.iloc[-1])
        v_prev = float(s.iloc[-2]) if len(s) >= 2 else None
        delta = (v_now - v_prev) if v_prev is not None else None
        summary_rows.append((col, v_now, delta))

    for i, (name, v, d) in enumerate(summary_rows, start=4):
        ws_sum[f"A{i}"] = name
        ws_sum[f"B{i}"] = v
        ws_sum[f"C{i}"] = d

    ws_sum["A14"] = "要因（メモ）"
    ws_sum["A15"] = "・（ここは後でLLMで自動生成して埋めるのが一番価値出る）"
    ws_sum["A17"] = "生成日時"
    ws_sum["B17"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    # --- Data
    export = master.copy().reset_index()
    first_col = export.columns[0]
    export = export.rename(columns={first_col: "month"})
    export["month"] = pd.to_datetime(export["month"]).dt.strftime("%Y-%m")
    _df_to_sheet(ws_data, export, start_row=1, start_col=1)

    # --- Charts (bytes -> BytesIO -> XLImage)
    chart_specs = []
    if "copper_jpy_kg" in master.columns:
        chart_specs.append(("copper_jpy_kg", "Copper (JPY/kg)", "JPY/kg"))
    if "aluminum_jpy_kg" in master.columns:
        chart_specs.append(("aluminum_jpy_kg", "Aluminum (JPY/kg)", "JPY/kg"))
   
    anchor_row = 1
    for col, title, ylab in chart_specs:
        s = master[col].dropna()
        if s.empty:
            continue
        png_bytes = _series_chart_png_bytes(s, title, ylab)
        img = XLImage(io.BytesIO(png_bytes))  # ★ファイル不要（TemporaryDirectory問題を回避）
        img.anchor = f"A{anchor_row}"
        ws_charts.add_image(img)
        anchor_row += 22  # 次の画像の位置

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

# ----------------------------
# 1) Copper / Aluminum (FRED)
# ----------------------------
copper = fetch_fred("PCOPPUSDM")
aluminum = fetch_fred("PALUMUSDM")
usdjpy = fetch_fred("EXJPUS")

df = pd.concat([copper, aluminum, usdjpy], axis=1, join="inner")
df.columns = ["copper_usd_ton", "aluminum_usd_ton", "usdjpy"]
df["copper_jpy_kg"] = df["copper_usd_ton"] * df["usdjpy"] / 1000
df["aluminum_jpy_kg"] = df["aluminum_usd_ton"] * df["usdjpy"] / 1000

latest_date = df.index[-1]
latest_month_str = latest_date.strftime("%Y-%m")

latest_copper = float(df.loc[latest_date, "copper_jpy_kg"])
latest_aluminum = float(df.loc[latest_date, "aluminum_jpy_kg"])

delta_copper = None
delta_aluminum = None
if len(df) >= 2:
    prev_date = df.index[-2]
    delta_copper = latest_copper - float(df.loc[prev_date, "copper_jpy_kg"])
    delta_aluminum = latest_aluminum - float(df.loc[prev_date, "aluminum_jpy_kg"])

st.subheader("📌 最新月の原価（円/kg）")
k1, k2, k3 = st.columns([1, 1, 1])
with k1:
    st.metric(
        label=f"銅（{latest_month_str}）",
        value=f"{latest_copper:,.0f} 円/kg",
        delta=f"{delta_copper:+,.0f} 円/kg" if delta_copper is not None else None,
    )
with k2:
    st.metric(
        label=f"アルミ（{latest_month_str}）",
        value=f"{latest_aluminum:,.0f} 円/kg",
        delta=f"{delta_aluminum:+,.0f} 円/kg" if delta_aluminum is not None else None,
    )
with k3:
    st.metric(label="更新日時", value=datetime.now().strftime("%Y-%m-%d %H:%M"))

tab1, tab2, tab3 = st.tabs(["🟠 銅", "⚪ アルミ", "📉 まとめ（同一グラフ）"])
with tab1:
    st.subheader("銅 価格推移（円/kg）")
    plot_with_latest_highlight(df["copper_jpy_kg"], "Copper (JPY/kg)", "JPY/kg")
with tab2:
    st.subheader("アルミ 価格推移（円/kg）")
    plot_with_latest_highlight(df["aluminum_jpy_kg"], "Aluminum (JPY/kg)", "JPY/kg")
with tab3:
    st.subheader("銅・アルミ（円/kg）同一グラフ")
    st.line_chart(df[["copper_jpy_kg", "aluminum_jpy_kg"]])

st.divider()

# ----------------------------
# 4) Excel export
# ----------------------------
st.divider()
st.subheader("📦 Excelレポート出力（表＋グラフ画像）")

try:
    df_fred = df[["copper_jpy_kg", "aluminum_jpy_kg"]].copy()
    master = build_monthly_master(df_fred)

    xlsx_bytes = make_excel_report(master)

    st.download_button(
        label="⬇️ Excelレポートをダウンロード",
        data=xlsx_bytes,
        file_name=f"cost_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.caption("※ ChartsシートにPNGグラフを添付し、Summary/Dataも同梱します。")

except Exception as e:
    st.error(f"Excel出力でエラー: {e}")




